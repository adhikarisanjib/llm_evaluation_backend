import json
from io import BytesIO
from pathlib import Path

import httpx
from aiosqlite import IntegrityError
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from pydantic import ValidationError
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.db import get_db
from app.experiment_runner import ExperimentRunner
from app.metrics import experiment_metrics
from app.models import Experiment, LLMModel, Run, Task
from app.schemas import (
    ExperimentCloneRequest,
    ExperimentCreate,
    ExperimentDetail,
    ExperimentRead,
    ExperimentUpdate,
    ModelCreate,
    ModelRead,
    ModelUpdate,
    RunRead,
    TaskCreate,
    TaskExcelUploadResult,
    TaskRead,
    TaskUpdate,
)
from app.settings import settings

router = APIRouter(prefix="/api")


HEADER_ALIASES = {
    "id": "id",
    "title": "title",
    "question": "question",
    "starter code": "starter_code",
    "starter_code": "starter_code",
    "difficulty": "difficulty",
    "task type": "task_type",
    "task_type": "task_type",
    "tests": "tests",
    "regression tests": "regression_tests",
    "regression_tests": "regression_tests",
    "programming language": "programming_language",
    "programming_language": "programming_language",
    "language": "programming_language",
}

REQUIRED_EXCEL_COLUMNS = {
    "title",
    "question",
    "task_type",
    "tests",
    "programming_language",
}


def _normalize_header(value) -> str:
    return " ".join(str(value or "").strip().lower().split())


def _parse_test_cell(value) -> list[str]:
    """Parse an Excel test cell.

    Friendly format: one assertion/test per line.
    Advanced format: a JSON array of strings for multi-line test snippets.
    """
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []

    if text.startswith("["):
        parsed = json.loads(text)
        if not isinstance(parsed, list) or not all(
            isinstance(item, str) for item in parsed
        ):
            raise ValueError("JSON test cells must be an array of strings")
        return [item.strip() for item in parsed if item.strip()]

    return [line.strip() for line in text.splitlines() if line.strip()]


def _optional_string(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


@router.get("/health")
async def health():
    return {"status": "ok"}


@router.get("/models/{model_id}", response_model=ModelRead)
async def get_model(model_id: int, db: AsyncSession = Depends(get_db)):
    model = await db.get(LLMModel, model_id)

    if not model:
        raise HTTPException(404, "Model not found")

    return model


@router.post("/models", response_model=ModelRead)
async def create_model(payload: ModelCreate, db: AsyncSession = Depends(get_db)):
    obj = LLMModel(**payload.model_dump())
    db.add(obj)
    await db.commit()
    await db.refresh(obj)
    return obj


@router.get("/models", response_model=list[ModelRead])
async def list_models(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(LLMModel).order_by(LLMModel.id))
    return result.scalars().all()


@router.patch("/models/{model_id}", response_model=ModelRead)
async def update_model(
    model_id: int, payload: ModelUpdate, db: AsyncSession = Depends(get_db)
):
    model = await db.get(LLMModel, model_id)

    if not model:
        raise HTTPException(404, "Model not found")

    update_data = payload.model_dump(exclude_unset=True)

    for field, value in update_data.items():
        setattr(model, field, value)

    try:
        await db.commit()
        await db.refresh(model)

    except Exception:
        await db.rollback()
        raise

    return model


@router.delete("/models/{model_id}", status_code=204)
async def delete_model(model_id: int, db: AsyncSession = Depends(get_db)):
    model = await db.get(LLMModel, model_id)

    if not model:
        raise HTTPException(404, "Model not found")

    run_count = await db.scalar(
        select(Run).where(Run.model_id == model_id).with_only_columns(func.count())
    )

    if run_count:
        raise HTTPException(
            status_code=409,
            detail=(
                f"Model cannot be deleted because it is referenced by {run_count} run(s). Disable the model instead if it should no longer be used."
            ),
        )

    try:
        await db.delete(model)
        await db.commit()

    except IntegrityError:
        await db.rollback()

        raise HTTPException(
            status_code=409,
            detail=(
                "Model cannot be deleted because it is referenced by an experiment."
            ),
        )

    return None


@router.post("/models/sync-ollama", response_model=list[ModelRead])
async def sync_ollama(
    base_url: str | None = Query(default=None),
    db: AsyncSession = Depends(get_db),
):
    ollama_url = (base_url or settings.ollama_base_url).rstrip("/")

    try:
        async with httpx.AsyncClient() as client:
            response = await client.get(f"{ollama_url}/api/tags", timeout=10)
            response.raise_for_status()
            items = response.json().get("models", [])
    except Exception as exc:
        raise HTTPException(502, f"Could not connect to Ollama: {exc}")

    result = await db.execute(
        select(LLMModel).where(
            LLMModel.adapter_type == "ollama",
            LLMModel.base_url == ollama_url,
        )
    )
    existing = {model.name: model for model in result.scalars().all()}

    for item in items:
        name = item["name"]
        if name not in existing:
            db.add(
                LLMModel(
                    name=name,
                    display_name=name,
                    adapter_type="ollama",
                    base_url=ollama_url,
                    metadata_json=item,
                )
            )
        else:
            existing[name].metadata_json = item
            existing[name].is_enabled = True

    await db.commit()

    result = await db.execute(
        select(LLMModel)
        .where(
            LLMModel.adapter_type == "ollama",
            LLMModel.base_url == ollama_url,
        )
        .order_by(LLMModel.id)
    )
    return result.scalars().all()


@router.post("/tasks", response_model=TaskRead)
async def create_task(payload: TaskCreate, db: AsyncSession = Depends(get_db)):
    data = payload.model_dump()
    task_id = data.pop("id", None)
    if task_id is not None:
        existing = await db.get(Task, task_id)
        if existing:
            raise HTTPException(409, f"Task ID {task_id} already exists")
        data["id"] = task_id

    task = Task(**data)
    db.add(task)
    await db.commit()
    await db.refresh(task)
    return task


@router.get("/tasks/excel-template")
async def download_task_excel_template():
    template = (
        Path(__file__).resolve().parents[2] / "examples" / "task_import_template.xlsx"
    )
    if not template.exists():
        raise HTTPException(404, "Task Excel template is not available")
    return FileResponse(
        template,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="task_import_template.xlsx",
    )


@router.post("/tasks/upload-excel", response_model=TaskExcelUploadResult)
async def upload_tasks_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """Import tasks from the first worksheet of an .xlsx file.

    Expected columns:
    id (optional), title, question, starter code, difficulty, task type, tests,
    regression tests, programming language.

    `tests` and `regression tests` accept either one test per Excel line or a
    JSON array of strings. The import is atomic: if any row is invalid, no rows
    are inserted.
    """
    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(400, "Only .xlsx Excel files are supported")

    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(413, "Excel file is too large; maximum size is 10 MB")
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
    except Exception as exc:
        raise HTTPException(400, f"Invalid Excel file: {exc}")

    rows = sheet.iter_rows(values_only=True)
    try:
        raw_headers = next(rows)
    except StopIteration:
        raise HTTPException(400, "Excel file is empty")

    column_map: dict[int, str] = {}
    unknown_headers: list[str] = []
    for index, value in enumerate(raw_headers):
        normalized = _normalize_header(value)
        if not normalized:
            continue
        mapped = HEADER_ALIASES.get(normalized)
        if mapped:
            column_map[index] = mapped
        else:
            unknown_headers.append(str(value))

    present = set(column_map.values())
    missing = sorted(REQUIRED_EXCEL_COLUMNS - present)
    if missing:
        raise HTTPException(
            422,
            {
                "message": "Missing required Excel columns",
                "missing_columns": missing,
                "unknown_columns": unknown_headers,
            },
        )

    parsed_tasks: list[TaskCreate] = []
    errors: list[dict] = []
    seen_ids: set[int] = set()
    last_task_id = await db.scalar(select(Task.id).order_by(Task.id.desc()).limit(1))
    if last_task_id is None:
        last_task_id = 0

    for excel_row_number, row in enumerate(rows, start=2):
        if not any(value is not None and str(value).strip() for value in row):
            continue

        raw = {
            field: row[index] if index < len(row) else None
            for index, field in column_map.items()
        }

        try:
            supplied_id = raw.get("id")
            if supplied_id is not None and str(supplied_id).strip():
                numeric_id = int(float(supplied_id)) + int(last_task_id)
                if numeric_id in seen_ids:
                    raise ValueError(
                        f"duplicate task id {numeric_id} inside Excel file"
                    )
                existing = await db.get(Task, numeric_id)
                if existing:
                    raise ValueError(f"task id {numeric_id} already exists in database")
                seen_ids.add(numeric_id)
            else:
                numeric_id = None

            payload = TaskCreate(
                id=numeric_id,
                title=str(raw.get("title") or "").strip(),
                question=str(raw.get("question") or "").strip(),
                starter_code=_optional_string(raw.get("starter_code")),
                difficulty=_optional_string(raw.get("difficulty")),
                task_type=str(raw.get("task_type") or "").strip(),
                tests=_parse_test_cell(raw.get("tests")),
                regression_tests=_parse_test_cell(raw.get("regression_tests")),
                programming_language=str(raw.get("programming_language") or "").strip(),
            )
            if not payload.title:
                raise ValueError("title is required")
            if not payload.question:
                raise ValueError("question is required")
            parsed_tasks.append(payload)
        except (ValueError, ValidationError, json.JSONDecodeError) as exc:
            errors.append({"row": excel_row_number, "error": str(exc)})

    if errors:
        raise HTTPException(
            422,
            {
                "message": "Excel import failed. No tasks were imported.",
                "errors": errors,
            },
        )

    if not parsed_tasks:
        raise HTTPException(400, "Excel file contains no task rows")

    created: list[Task] = []
    try:
        for payload in parsed_tasks:
            data = payload.model_dump()
            task_id = data.pop("id", None)
            if task_id is not None:
                data["id"] = task_id
            task = Task(**data)
            db.add(task)
            created.append(task)
        await db.commit()
        for task in created:
            await db.refresh(task)
    except Exception:
        await db.rollback()
        raise

    return TaskExcelUploadResult(
        imported_count=len(created),
        task_ids=[task.id for task in created],
    )


@router.get("/tasks", response_model=list[TaskRead])
async def list_tasks(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Task).order_by(Task.id))
    return result.scalars().all()


@router.get("/tasks/{task_id}", response_model=TaskRead)
async def get_task(task_id: int, db: AsyncSession = Depends(get_db)):
    task = await db.get(Task, task_id)
    if not task:
        raise HTTPException(404, "Task not found")
    return task


@router.patch("/tasks/{task_id}", response_model=TaskRead)
async def update_task(
    task_id: int, payload: TaskUpdate, db: AsyncSession = Depends(get_db)
):
    task = await db.get(Task, task_id)

    if not task:
        raise HTTPException(404, "Task not found")

    update_data = payload.model_dump(exclude_unset=True)

    for field, value in update_data.items():
        setattr(task, field, value)

    try:
        await db.commit()
        await db.refresh(task)

    except Exception:
        await db.rollback()
        raise

    return task


@router.delete("/tasks/{task_id}", status_code=204)
async def delete_task(task_id: int, db: AsyncSession = Depends(get_db)):
    task = await db.get(Task, task_id)

    if not task:
        raise HTTPException(404, "Task not found")

    run_count = await db.scalar(
        select(func.count()).select_from(Run).where(Run.task_id == task_id)
    )

    if run_count:
        raise HTTPException(
            status_code=409,
            detail=(
                f"Task cannot be deleted because it is referenced by {run_count} run(s). Deleting it would invalidate existing experiment results."
            ),
        )

    try:
        await db.delete(task)
        await db.commit()

    except IntegrityError:
        await db.rollback()

        raise HTTPException(
            status_code=409,
            detail=(
                "Task cannot be deleted because it is currently referenced by one or more experiments."
            ),
        )

    return None


@router.post("/experiments", response_model=ExperimentRead)
async def create_experiment(
    payload: ExperimentCreate, db: AsyncSession = Depends(get_db)
):
    result = await db.execute(
        select(LLMModel).where(LLMModel.id.in_(payload.model_ids))
    )
    models = result.scalars().all()

    result = await db.execute(select(Task).where(Task.id.in_(payload.task_ids)))
    tasks = result.scalars().all()

    if len(models) != len(set(payload.model_ids)):
        raise HTTPException(400, "One or more model IDs do not exist")
    if len(tasks) != len(set(payload.task_ids)):
        raise HTTPException(400, "One or more task IDs do not exist")

    disabled = [model.id for model in models if not model.is_enabled]
    if disabled:
        raise HTTPException(400, f"Disabled models cannot be selected: {disabled}")

    experiment = Experiment(
        name=payload.name,
        attempts_per_task=payload.attempts_per_task,
        temperature=payload.temperature,
        top_p=payload.top_p,
        max_tokens=payload.max_tokens,
    )
    experiment.models = models
    experiment.tasks = tasks
    db.add(experiment)
    await db.commit()
    await db.refresh(experiment)
    return experiment


@router.get("/experiments", response_model=list[ExperimentRead])
async def list_experiments(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Experiment).order_by(Experiment.id.desc()))
    return result.scalars().all()


@router.get("/experiments/{experiment_id}", response_model=ExperimentDetail)
async def get_experiment(experiment_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Experiment)
        .options(
            selectinload(Experiment.models),
            selectinload(Experiment.tasks),
        )
        .where(Experiment.id == experiment_id)
    )
    experiment = result.scalar_one_or_none()
    if not experiment:
        raise HTTPException(404, "Experiment not found")
    return experiment


@router.patch("/experiments/{experiment_id}", response_model=ExperimentDetail)
async def update_experiment(
    experiment_id: int, payload: ExperimentUpdate, db: AsyncSession = Depends(get_db)
):
    result = await db.execute(
        select(Experiment)
        .options(
            selectinload(Experiment.models),
            selectinload(Experiment.tasks),
        )
        .where(Experiment.id == experiment_id)
    )

    experiment = result.scalar_one_or_none()

    if not experiment:
        raise HTTPException(404, "Experiment not found")

    update_data = payload.model_dump(exclude_unset=True)

    model_ids = update_data.pop("model_ids", None)

    task_ids = update_data.pop("task_ids", None)

    for field, value in update_data.items():
        setattr(experiment, field, value)

    if model_ids is not None:

        if not model_ids:
            raise HTTPException(400, "Experiment must contain at least one model")

        result = await db.execute(select(LLMModel).where(LLMModel.id.in_(model_ids)))

        models = result.scalars().all()

        if len(models) != len(set(model_ids)):
            raise HTTPException(400, "One or more model IDs do not exist")

        disabled = [model.id for model in models if not model.is_enabled]

        if disabled:
            raise HTTPException(
                400,
                f"Disabled models cannot be selected: {disabled}",
            )

        experiment.models = models

    if task_ids is not None:

        if not task_ids:
            raise HTTPException(400, "Experiment must contain at least one task")

        result = await db.execute(select(Task).where(Task.id.in_(task_ids)))

        tasks = result.scalars().all()

        if len(tasks) != len(set(task_ids)):
            raise HTTPException(400, "One or more task IDs do not exist")

        experiment.tasks = tasks

    try:
        await db.commit()

    except Exception:
        await db.rollback()
        raise

    # Reload relationships after commit
    result = await db.execute(
        select(Experiment)
        .options(
            selectinload(Experiment.models),
            selectinload(Experiment.tasks),
        )
        .where(Experiment.id == experiment_id)
    )

    return result.scalar_one()


@router.delete("/experiments/{experiment_id}", status_code=204)
async def delete_experiment(experiment_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Experiment)
        .options(
            selectinload(Experiment.models),
            selectinload(Experiment.tasks),
        )
        .where(Experiment.id == experiment_id)
    )

    experiment = result.scalar_one_or_none()

    if not experiment:
        raise HTTPException(404, "Experiment not found")

    try:
        await db.delete(experiment)
        await db.commit()

    except IntegrityError as exc:
        await db.rollback()

        raise HTTPException(
            status_code=409,
            detail=(
                "Experiment could not be deleted. Check that experiment runs are configured with delete cascade."
            ),
        ) from exc

    return None


@router.post("/experiments/{experiment_id}/clone", response_model=ExperimentRead)
async def clone_experiment(
    experiment_id: int,
    payload: ExperimentCloneRequest | None = None,
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(Experiment)
        .options(
            selectinload(Experiment.models),
            selectinload(Experiment.tasks),
        )
        .where(Experiment.id == experiment_id)
    )
    source = result.scalar_one_or_none()
    if not source:
        raise HTTPException(404, "Experiment not found")

    payload = payload or ExperimentCloneRequest()
    clone = Experiment(
        name=payload.name or f"{source.name} (repeat)",
        attempts_per_task=(
            payload.attempts_per_task
            if payload.attempts_per_task is not None
            else source.attempts_per_task
        ),
        temperature=source.temperature,
        top_p=source.top_p,
        max_tokens=source.max_tokens,
    )
    clone.models = list(source.models)
    clone.tasks = list(source.tasks)
    db.add(clone)
    await db.commit()
    await db.refresh(clone)
    return clone


@router.post("/experiments/{experiment_id}/run", response_model=ExperimentRead)
async def run_experiment(experiment_id: int, db: AsyncSession = Depends(get_db)):
    experiment = await db.get(Experiment, experiment_id)
    if not experiment:
        raise HTTPException(404, "Experiment not found")
    return await ExperimentRunner(db).run(experiment_id)


@router.post("/experiments/{experiment_id}/retry-failed", response_model=ExperimentRead)
async def retry_failed_experiment(
    experiment_id: int, db: AsyncSession = Depends(get_db)
):
    experiment = await db.get(Experiment, experiment_id)
    if not experiment:
        raise HTTPException(404, "Experiment not found")
    return await ExperimentRunner(db).retry_failed(experiment_id)


@router.get("/experiments/{experiment_id}/runs", response_model=list[RunRead])
async def experiment_runs(experiment_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Run)
        .options(selectinload(Run.test_results))
        .where(Run.experiment_id == experiment_id)
        .order_by(Run.id)
    )
    return result.scalars().all()


@router.get("/runs/{run_id}", response_model=RunRead)
async def get_run(run_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Run).options(selectinload(Run.test_results)).where(Run.id == run_id)
    )
    run = result.scalar_one_or_none()
    if not run:
        raise HTTPException(404, "Run not found")
    return run


@router.get("/experiments/{experiment_id}/metrics")
async def metrics(experiment_id: int, db: AsyncSession = Depends(get_db)):
    experiment = await db.get(Experiment, experiment_id)
    if not experiment:
        raise HTTPException(404, "Experiment not found")
    return await experiment_metrics(db, experiment_id)
